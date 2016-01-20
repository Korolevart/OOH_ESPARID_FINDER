
import xlrd
import openpyxl
import numpy
import math
import openpyxl
import xlrd
import urllib
import urllib.request
import re
from openpyxl import load_workbook
from time import sleep
import pandas as pd
import random

#Checking dictionaries and replacing names in working_file
def dictionary_check(working_file, dictionary_add, name):
    #Reading City_index
    dictionary_add = xlrd.open_workbook(dictionary_add)
    sheet = dictionary_add.sheet_by_name('Sheet1')
    # read City Index from Excel and replace cities in EXCEL and SQL
    keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]
    dict_list = []
    for row_index in range(1, sheet.nrows):
        d = {keys[col_index]: sheet.cell(row_index, col_index).value for col_index in range(sheet.ncols)}
        a=d.get(name)
        b=d.get('abc')
        working_file[name]=working_file[name].replace(a, b)
        #df_SQL['CITY']=df_SQL['CITY'].replace(a, b)
        dict_list.append(d)
    return working_file

def location(df_SQL,ESP,hashdata,lat,lng,Owner,Sizes,ABC,City):
    hashdata=[]
    hashdata=ESP
    a=0
    hashdata[lat] = numpy.nan
    hashdata[lng] = numpy.nan
    hashdata[Owner] = ''
    hashdata[Sizes] = ''
    hashdata[ABC] = ''
    hashdata[City] = ''
    for espar_id in hashdata['ESPAR_ID']:
        print(espar_id)

        print(a)
        b=0
        for row in df_SQL['ESPAR_ID']:
            #print(row)
            #print(b)
            if row == espar_id:
                hashdata.set_value(a, lng, df_SQL.iloc[b]['X1'])
                hashdata.set_value(a, lat, df_SQL.iloc[b]['Y1'])
                hashdata.set_value(a, Owner, df_SQL.iloc[b]['OWNER'])
                hashdata.set_value(a, Sizes, df_SQL.iloc[b]['SIZES'])
                hashdata.set_value(a, ABC, df_SQL.iloc[b]['ABC'])
                hashdata.set_value(a, City, df_SQL.iloc[b]['CITY3'])
                hashdata.set_value(a, City, df_SQL.iloc[b]['CITY3'])
            b+=1
        a+=1
    return hashdata


#Looking for ESPAR_ids using KEYs and coordinates
def KEY_loop(df_working, df_SQL, KEY, Distan, ESPAR_IDs):
    if df_working.shape[0]>0:
        #Running loop with KEY1
        print(KEY)
        a=0
        df_working[Distan] = numpy.nan
        df_working[ESPAR_IDs] = ''
        for all_SQL_rows in df_SQL[KEY]:
            b=0
            for all_working_rows in df_working[KEY]:
                if (all_SQL_rows == all_working_rows and math.isnan(df_working.iloc[b]['lat']) != True):
                    Distance=numpy.sqrt((df_SQL.iloc[a]['Y1']-df_working.iloc[b]['lat'])**2+(df_SQL.iloc[a]['X1']-df_working.iloc[b]['lng'])**2)
                    if (math.isnan(df_working[Distan][b]) and df_SQL.iloc[a]['ESPAR_ID'] not in df_working[ESPAR_IDs]):
                        df_working.set_value(b, Distan, Distance)
                        df_working.set_value(b, ESPAR_IDs, df_SQL.iloc[a]['ESPAR_ID'])
                    elif (df_working[Distan][b]>Distance and df_SQL.iloc[a]['ESPAR_ID'] not in df_working[ESPAR_IDs]):
                        df_working.set_value(b, Distan, Distance)
                        df_working.set_value(b, ESPAR_IDs, df_SQL.iloc[a]['ESPAR_ID'])
                b=b+1
            a=a+1
        return df_working



#searching for address in google and yandex and retrieve coordinates
def Google_search(df_working):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import re
    df_working['Search'] =df_working['City'] +", " + df_working['Adress2']
    df_working['lat'] = numpy.nan
    df_working['lng'] = numpy.nan
    a=0
    for i in df_working['Search']:
        print(a)
        url = 'https://geocode-maps.yandex.ru/1.x/?geocode='
        word=str(i)
        #word=df_working['Search'][6]
        print(word)
        if (('МКАД' in word) or ('км' in word) or ('километр' in word) or ('/' in word)): #YANDEX
            service_args = ['--proxy=mskwebp01101:8080','--proxy-auth=Artem.Korolev@mecglobal.com:WertelS3boge1',]
            driver = webdriver.PhantomJS('C:/Python34/phantomjs-2.0.0-windows/bin/phantomjs.exe', service_args=service_args)
            #driver.set_window_size(1120, 550)
            driver.get('https://maps.yandex.ru/')
            #sleep(random.uniform(1, 5))
            try:
                element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/header/div[1]/div/form/div[2]/button")) and EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/header/div[1]/div/form/div[1]/div/span/span/input"))
                )
            finally:
                search = driver.find_element_by_xpath(("/html/body/div[1]/header/div[1]/div/form/div[1]/div/span/span/input"))
                button = driver.find_element_by_xpath(("/html/body/div[1]/header/div[1]/div/form/div[2]/button"))
                print('TRUE')
                print(word)
                search.send_keys(word)
                #sleep(random.uniform(1, 5))
                button.submit()
                #sleep(random.uniform(1, 5))
                driver.save_screenshot('out0.png')
                #sleep(4)
            try:
                element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div/div[1]/div/div[2]/div[1]/div/div[1]/div[1]/div[2]")) and EC.presence_of_element_located((By.XPATH,"/html/body/div[1]/div[2]/div/div[1]/div/div[2]/div[1]/div/div[1]/div[2]/div[2]"))
                )
            finally:
                lan=driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[1]/div/div[2]/div[1]/div/div[1]/div[1]/div[2]").text
                driver.save_screenshot('out1.png')
                lng=driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[1]/div/div[2]/div[1]/div/div[1]/div[2]/div[2]").text
                driver.save_screenshot('out2.png')
                lat = re.search('([0-9]{1,}[,.][0-9]{1,})', lan)
                lat=lat.group(0)
                lng = re.search('([0-9]{1,}[,.][0-9]{1,})', lng)
                lng=lng.group(0)
            driver.quit()
        else: #Yandex
            full_url = url+word
            print(full_url)
            if full_url != 'https://geocode-maps.yandex.ru/1.x/?geocode=nan':
                url = urllib.parse.urlsplit(full_url)
                #print(url)
                russian = ''.join(re.findall('[А-Яа-я,.\/\\-\d]', url[3]))
                #print(russian)
                urlnew = urllib.parse.quote(russian)
                url_last=url[0]+"://"+url[1]+url[2]+"?geocode="+urlnew
                proxy_support = urllib.request.ProxyHandler({})
                opener = urllib.request.build_opener(proxy_support)
                urllib.request.install_opener(opener)
                with urllib.request.urlopen(url_last) as response:
                    html = response.read()
                address = re.search('(<lowerCorner>)([0-9]{1,}[,.][0-9]{1,})( )([0-9]{1,}[,.][0-9]{1,})', str(html))
                lat = re.search('([0-9]{1,}[,.][0-9]{1,})(</pos>)', str(html))
                lng = re.search('(<pos>)([0-9]{1,}[,.][0-9]{1,})', str(html))
                print(word+' '+'address: ' +' '+'lat: '+lat.group(1)+' '+'lng: '+lng.group(2))
                lat = lat.group(1)
                lng = lng.group(2)
        df_working.set_value(a, 'lat', lat)
        df_working.set_value(a, 'lng', lng)
        a=a+1
        sleep(0.3)
    return df_working

def KEY_loop_replacing_uniq(df_working, df_SQL, KEY, Distan, ESPAR_IDs, Distance):
    print('duplicated')
    df_working[Distan] = ''
    df_working=df_working.sort([ESPAR_IDs])
    df_working=df_working[df_working[ESPAR_IDs]!='']
    df_working.reset_index(drop=True,inplace =True)
    counter=df_working[ESPAR_IDs].count()
    b=0
    for b in range(1,counter-1):
        word1=df_working[ESPAR_IDs][b]
        word2=df_working[ESPAR_IDs][b+1]
        if (df_working[ESPAR_IDs][b]==df_working[ESPAR_IDs][b+1]):
            a=0
            for all_SQL_rows in df_SQL[KEY]:
                DistanceNew=numpy.sqrt((df_SQL.iloc[a]['Y1']-df_working['lat'][b])**2+(df_SQL.iloc[a]['X1']-df_working['lng'][b])**2)
                if ((all_SQL_rows == df_working[KEY][b]) and (DistanceNew < Distance) and (df_SQL.iloc[a]['ESPAR_ID'] not in df_working[ESPAR_IDs].unique().tolist())):
                    df_working.set_value(b, Distan, DistanceNew)
                    df_working.set_value(b, ESPAR_IDs, df_SQL.iloc[a]['ESPAR_ID'])
                    print(DistanceNew)
                else:
                    df_working.set_value(b, ESPAR_IDs, 'MANUAL')
                a=a+1
        b=b+1

    return df_working
