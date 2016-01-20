import sys
import numpy
import math
import pandas as pd
from PyQt4 import QtGui, QtCore
#Checking dictionaries and replacing names in working_file
def dictionary_check(working_file, dictionary_add, name):
    import xlrd
    from openpyxl import load_workbook
    #Reading City_index
    dictionary_add = xlrd.open_workbook(dictionary_add)
    sheet = dictionary_add.sheet_by_name('Sheet1')
    # read City Index from Excel and replace cities in EXCEL and SQL
    keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]
    dict_list = []
    for row_index in range(1, sheet.nrows):
        d = {keys[col_index]: sheet.cell(row_index, col_index).value for col_index in range(sheet.ncols)}
        a=d.get(name)
        b=d.get('abc')
        working_file[name]=working_file[name].replace(a, b)
        #df_SQL['CITY']=df_SQL['CITY'].replace(a, b)
        dict_list.append(d)
    return working_file

    
def KEY_loop_replacing_uniq(df_working, df_SQL, KEY, Distan, ESPAR_IDs, Distance):
    print('duplicated')
    df_working=df_working.sort([ESPAR_IDs])
    df_working=df_working[df_working[ESPAR_IDs]!='']
    df_working.reset_index(drop=True,inplace =True)
    counter=df_working[ESPAR_IDs].count()
    b=0
    for b in range(1,counter-1):
        word1=df_working[ESPAR_IDs][b]
        word2=df_working[ESPAR_IDs][b+1]
        if (df_working[ESPAR_IDs][b]==df_working[ESPAR_IDs][b+1]):
            a=0
            for all_SQL_rows in df_SQL[KEY]:
                DistanceNew=numpy.sqrt((df_SQL.iloc[a]['Y1']-df_working['lat'][b])**2+(df_SQL.iloc[a]['X1']-df_working['lng'][b])**2)
                if ((all_SQL_rows == df_working[KEY][b]) and ((DistanceNew-df_working.iloc[b][Distan]) < Distance) and (df_SQL.iloc[a]['ESPAR_ID'] not in df_working[ESPAR_IDs].unique().tolist())):
                    df_working.set_value(b, Distan, DistanceNew)
                    df_working.set_value(b, ESPAR_IDs, df_SQL.iloc[a]['ESPAR_ID'])
                    print(DistanceNew)
                else:
                    df_working.set_value(b, ESPAR_IDs, 'MANUAL')
                a=a+1
        b=b+1

    return df_working

    
    #Looking for ESPAR_ids using KEYs and coordinates
def KEY_loop(df_working, df_SQL, KEY, Distan, ESPAR_IDs):
    #Running loop with KEY1
    print(KEY)
    a=0
    df_working[Distan] = numpy.nan
    df_working[ESPAR_IDs] = ''
    for all_SQL_rows in df_SQL[KEY]:
        b=0
        for all_working_rows in df_working[KEY]:
            if (all_SQL_rows == all_working_rows and math.isnan(df_working.iloc[b]['lat']) != True):
                Distance=numpy.sqrt((df_SQL.iloc[a]['Y1']-df_working.iloc[b]['lat'])**2+(df_SQL.iloc[a]['X1']-df_working.iloc[b]['lng'])**2)
                if (math.isnan(df_working[Distan][b]) and df_SQL.iloc[a]['ESPAR_ID'] not in df_working[ESPAR_IDs]):
                    df_working.set_value(b, Distan, Distance)
                    df_working.set_value(b, ESPAR_IDs, df_SQL.iloc[a]['ESPAR_ID'])
                elif (df_working[Distan][b]>Distance and df_SQL.iloc[a]['ESPAR_ID'] not in df_working[ESPAR_IDs]):
                    df_working.set_value(b, Distan, Distance)
                    df_working.set_value(b, ESPAR_IDs, df_SQL.iloc[a]['ESPAR_ID'])
            b=b+1
        a=a+1
    return df_working
    

class Window(QtGui.QMainWindow):

    def __init__(self):
        super(Window, self).__init__()

        self.setGeometry(50, 50, 700, 300)
        self.setWindowTitle("OOH")
        self.setWindowIcon(QtGui.QIcon('pythonlogo.png'))

        extractAction = QtGui.QAction("&Quiet", self)
        extractAction.setShortcut("Ctrl+Q")
        extractAction.setStatusTip('Leave The App')
        extractAction.triggered.connect(self.close_application)

        openFile = QtGui.QAction("&Open File", self)
        openFile.setShortcut("Ctrl+O")
        openFile.setStatusTip('Open File')
        openFile.triggered.connect(self.file_open)

        saveFile = QtGui.QAction("&Save File", self)
        saveFile.setShortcut("Ctrl+S")
        saveFile.setStatusTip('Save File')
        saveFile.triggered.connect(self.file_save)

        self.statusBar()

        
        #editorMenu = mainMenu.addMenu("&Editor")
        #editorMenu.addAction(openEditor)

        data_new = []
        region = list()
        SQL = []
        
        self.home()
        
    def home(self):
        btn = QtGui.QPushButton("Quit", self)
        btn.clicked.connect(self.close_application)
        btn.resize(btn.minimumSizeHint())
        btn.move(0,100)

        self.text_edit = QtGui.QTextEdit(self)
        self.setCentralWidget(self.text_edit)
        

        extractAction = QtGui.QAction(QtGui.QIcon('C:/Users/artem.korolev/Desktop/load.png'), 'Load Excel File', self)
        extractAction.triggered.connect(self.file_open)
        self.toolBar = self.addToolBar("Load")
        self.toolBar.addAction(extractAction)

        extractAction = QtGui.QAction(QtGui.QIcon('C:/Users/artem.korolev/Desktop/search.png'), 'Search for location', self)
        extractAction.triggered.connect(self.google_search)
        self.toolBar = self.addToolBar("Search")
        self.toolBar.addAction(extractAction)

        extractAction = QtGui.QAction(QtGui.QIcon('C:/Users/artem.korolev/Desktop/index.png'), 'Loadinf SQL', self)
        extractAction.triggered.connect(self.file_SQL)
        self.toolBar = self.addToolBar("Search")
        self.toolBar.addAction(extractAction)

        extractAction = QtGui.QAction(QtGui.QIcon('C:/Users/artem.korolev/Desktop/billboard'), 'Look for Espar ID', self)
        extractAction.triggered.connect(self.main_function)
        self.toolBar = self.addToolBar("Search Espar ID")
        self.toolBar.addAction(extractAction)

        extractAction = QtGui.QAction(QtGui.QIcon('C:/Users/artem.korolev/Desktop/save.png'), 'Save current Excel file', self)
        extractAction.triggered.connect(self.file_save)
        self.toolBar = self.addToolBar("Save")
        self.toolBar.addAction(extractAction)

        extractAction = QtGui.QAction(QtGui.QIcon('C:/Users/artem.korolev/Desktop/close.png'), 'Close application', self)
        extractAction.triggered.connect(self.close_application)
        self.toolBar = self.addToolBar("Quiet")
        self.toolBar.addAction(extractAction)

        self.progress = QtGui.QProgressBar(self)
        self.progress.setGeometry(200, 100, 250, 30)
        self.progress.move(400,23)
        
        self.comboBox = QtGui.QComboBox(self)
        self.comboBox.move(270,23)
        self.comboBox.addItem('AST')
        self.comboBox.addItem('BAR')
        self.comboBox.addItem('BLG')
        self.comboBox.addItem('BRN')
        self.comboBox.addItem('CHL')
        self.comboBox.addItem('CHP')
        self.comboBox.addItem('EKT')
        self.comboBox.addItem('HAB')
        self.comboBox.addItem('IRK')
        self.comboBox.addItem('IZH')
        self.comboBox.addItem('KLG')
        self.comboBox.addItem('KRA')
        self.comboBox.addItem('KRD')
        self.comboBox.addItem('KRV')
        self.comboBox.addItem('KUR')
        self.comboBox.addItem('KZN')
        self.comboBox.addItem('LPK')
        self.comboBox.addItem('MOS')
        self.comboBox.addItem('NNG')
        self.comboBox.addItem('NVG')
        self.comboBox.addItem('NVS')
        self.comboBox.addItem('OMS')
        self.comboBox.addItem('ORL')
        self.comboBox.addItem('PRM')
        self.comboBox.addItem('RND')
        self.comboBox.addItem('RZN')
        self.comboBox.addItem('SML')
        self.comboBox.addItem('SMR')
        self.comboBox.addItem('SPB')
        self.comboBox.addItem('SRT')
        self.comboBox.addItem('TLT')
        self.comboBox.addItem('TMS')
        self.comboBox.addItem('TUL')
        self.comboBox.addItem('TUM')
        self.comboBox.addItem('UFA')
        self.comboBox.addItem('ULV')
        self.comboBox.addItem('VDV')
        self.comboBox.addItem('VLD')
        self.comboBox.addItem('VLG')
        self.comboBox.addItem('VOR')
        self.comboBox.addItem('YAR')
        
        self.show()

    def file_open(self):
        import pandas as pd
        name = QtGui.QFileDialog.getOpenFileName(self, 'Open File')
        data_working = pd.read_excel(name,'Sheet1',index_col=None, na_values=['NA'])
        self.text_edit.setText('Loadin Excel file...')
        self.data_new = data_working
        self.text_edit.append('File is loaded')
        print(self.data_new)

    def file_SQL(self):
        import sqlalchemy
        import pyodbc
        import pandas as pd
        self.region=str(self.comboBox.currentText())
        self.text_edit.append("Getting data from SQL")
        print("Getting data from SQL")
        print(self.region)
        engine = sqlalchemy.create_engine("mssql+pyodbc://espar:espar@MSKSQLP01110/Odaplan?driver=SQL Server")
        request="""SELECT  distinct b.[CITY3],a.[ESPAR_ID],a.[OWNER], b.[X1], b.[Y1],b.[TYPE_GENER], b.[SIZES],b.[ABC]
                FROM [Odaplan].[dbo].[OP] a join [Odaplan].[dbo].[OSDATA] b on a.[ESPAR_ID]=b.[ESPAR_ID] where a.[PROJECT_ID]=(SELECT [PROJECT_ID]
                FROM [Odaplan].[dbo].[PROJECTS] where [NAME]='2015_1' ) and b.X1 is not null and b.[CITY3]="""+"'"+self.region+"'"
        self.SQL=pd.read_sql(request,engine)
        print(self.SQL)
        self.text_edit.append("DONE")

    def file_save(self):
        new_data=self.data_new
        import pandas as pd
        name = QtGui.QFileDialog.getSaveFileName(self, 'Save File')
        writer = pd.ExcelWriter(str(name)+str('.xlsx'))
        new_data.to_excel(writer,'Sheet1')
        writer.save()


    def editor(self):
        self.textEdit = QtGui.QTextEdit()
        self.setCentralWidget(self.textEdit)


    def google_search(self):
        import xlrd
        import openpyxl
        import numpy
        import math
        import openpyxl
        import xlrd
        import urllib
        import urllib.request
        import re
        from openpyxl import load_workbook
        from time import sleep
        import pandas as pd
        import random
        df_working=self.data_new
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import re
        df_working['Search'] =df_working['City'] +", " + df_working['Adress2']
        df_working['lat'] = numpy.nan
        df_working['lng'] = numpy.nan
        a=0
        self.completed = 0
        Counter = df_working.shape[0]
        Step = 100/Counter
        for i in df_working['Search']:
            self.completed += Step
            self.progress.setValue(self.completed)
            self.text_edit.append(str(a))
            print(str(a))
            url = 'https://geocode-maps.yandex.ru/1.x/?geocode='
            word=str(i)
            #word=df_working['Search'][70]
            self.text_edit.append(word)
            print(word)
            if (('МКАД' in word) or ('км' in word) or ('километр' in word) or ('/' in word)): #YANDEX
                service_args = ['--proxy=mskwebp01101:8080','--proxy-auth=Artem.Korolev@mecglobal.com:WertelS3boge1',]
                driver = webdriver.PhantomJS('C:/Python34/phantomjs-2.0.0-windows/bin/phantomjs.exe', service_args=service_args)
                #driver.set_window_size(1120, 550)
                driver.get('https://maps.yandex.ru/')
                #sleep(random.uniform(1, 5))
                try:
                    element = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/header/div[1]/div/form/div[2]/button")) and EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/header/div[1]/div/form/div[1]/div/span/span/input"))
                    )
                finally:
                    search = driver.find_element_by_xpath(("/html/body/div[1]/header/div[1]/div/form/div[1]/div/span/span/input"))
                    button = driver.find_element_by_xpath(("/html/body/div[1]/header/div[1]/div/form/div[2]/button"))
                    self.text_edit.append(word)
                    print('TRUE')
                    print(word)
                    search.send_keys(word)
                    #sleep(random.uniform(1, 5))
                    button.submit()
                    #sleep(random.uniform(1, 5))
                    driver.save_screenshot('out0.png')
                    #sleep(4)
                try:
                    element = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div/div[1]/div/div[2]/div[1]/div/div[1]/div[1]/div[2]")) and EC.presence_of_element_located((By.XPATH,"/html/body/div[1]/div[2]/div/div[1]/div/div[2]/div[1]/div/div[1]/div[2]/div[2]"))
                    )
                finally:
                    lan=driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[1]/div/div[2]/div[1]/div/div[1]/div[1]/div[2]").text
                    driver.save_screenshot('out1.png')
                    lng=driver.find_element_by_xpath("/html/body/div[1]/div[2]/div/div[1]/div/div[2]/div[1]/div/div[1]/div[2]/div[2]").text
                    driver.save_screenshot('out2.png')
                    lat = re.search('([0-9]{1,}[,.][0-9]{1,})', lan)
                    lat=lat.group(0)
                    lng = re.search('([0-9]{1,}[,.][0-9]{1,})', lng)
                    lng=lng.group(0)
                    driver.quit()
            else: #Yandex
                full_url = url+word
                print(full_url)
                self.text_edit.append(full_url)
                if full_url != 'http://geocode-maps.yandex.ru/1.x/?geocode=nan':
                    url = urllib.parse.urlsplit(full_url)
                    #print(url)
                    russian = ''.join(re.findall('[А-Яа-я,.\/\\-\d]', url[3]))
                    #print(russian)
                    urlnew = urllib.parse.quote(russian)
                    url_last=url[0]+"://"+url[1]+url[2]+"?geocode="+urlnew
                    proxy_support = urllib.request.ProxyHandler({})
                    opener = urllib.request.build_opener(proxy_support)
                    urllib.request.install_opener(opener)
                    with urllib.request.urlopen(url_last) as response:
                        html = response.read()
                    address = re.search('(<lowerCorner>)([0-9]{1,}[,.][0-9]{1,})( )([0-9]{1,}[,.][0-9]{1,})', str(html))
                    lat = re.search('([0-9]{1,}[,.][0-9]{1,})(</pos>)', str(html))
                    lng = re.search('(<pos>)([0-9]{1,}[,.][0-9]{1,})', str(html))
                    print(word+' '+'address: ' +' '+'lat: '+lat.group(1)+' '+'lng: '+lng.group(2))
                    self.text_edit.append(word+' '+'address: ' +' '+'lat: '+lat.group(1)+' '+'lng: '+lng.group(2))
                    lat = lat.group(1)
                    lng = lng.group(2)
            df_working.set_value(a, 'lat', lat)
            df_working.set_value(a, 'lng', lng)
            a=a+1
            sleep(0.3)
        self.data_new=df_working


        

    def close_application(self):
        choice = QtGui.QMessageBox.question(self, 'Extract!',
                                            "Get into the chopper?",
                                            QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
        if choice == QtGui.QMessageBox.Yes:
            print("Extracting Naaaaaaoooww!!!!")
            sys.exit()
        else:
            pass

    
    def main_function(self):
        #Reading City_index and making changes
        df_working = self.data_new
        df_SQL = self.SQL
        print("Reading City_index and making changes")
        df_working = dictionary_check(df_working,'J:/MEC/Analytics and Insight/Korolev/Python/OOH/CITY_index.xlsx','City')
        #Reading ABC_index and making changes
        print("Reading ABC_index and making changes")
        df_working = dictionary_check(df_working,'J:/MEC/Analytics and Insight/Korolev/Python/OOH/ABC_index.xlsx','ABC')
        #Reading Owner_index and making changes
        print("Reading Owner_index and making changes")
        df_working = dictionary_check(df_working,'J:/MEC/Analytics and Insight/Korolev/Python/OOH/OWNER_index.xlsx','Owner')
        #Reading Sizes_index and making changes
        print("Reading Sizes_index and making changes")
        df_working = dictionary_check(df_working,'J:/MEC/Analytics and Insight/Korolev/Python/OOH/SIZES_index.xlsx','Sizes')

        #Creating Key columns in data from group
        #df_working['CITY'] + df_working['OWNER']+ df_working['SIZES']+ df_working['ABC']
        df_working['Key1'] = df_working['City'] + df_working['Owner']+ df_working['Sizes']+ df_working['ABC']
        #df_working['CITY'] + df_working['SIZES']+ df_working['ABC']
        df_working['Key2'] = df_working['City'] + df_working['Sizes']+ df_working['ABC']
        #df_working['CITY'] + df_working['SIZES']
        df_working['Key3'] = df_working['City'] + df_working['Sizes']


        #Creating Key column in data from SQL
        #df_SQL['CITY'] + df_SQL['OWNER']+ df_SQL['TYPE_GENER']+ df_SQL['SIZES']+ df_SQL['ABC']
        df_SQL['Key1'] = df_SQL['CITY3'] + df_SQL['OWNER']+ df_SQL['SIZES']+ df_SQL['ABC']
        #df_SQL['CITY'] + df_SQL['TYPE_GENER']+ df_SQL['SIZES']+ df_SQL['ABC']
        df_SQL['Key2'] = df_SQL['CITY3'] + df_SQL['SIZES']+ df_SQL['ABC']
        #df_SQL['CITY'] + df_SQL['TYPE_GENER']+ df_SQL['SIZES']
        df_SQL['Key3'] = df_SQL['CITY3'] + df_SQL['SIZES']

        #Running loop with KEY1
        df_working=KEY_loop(df_working, df_SQL, 'Key1', 'Distan1', 'ESPAR_ID1')
        #duplicates=KEY_loop_replacing_uniq(df_working, df_SQL, 'Key1', 'Distan1', 'ESPAR_ID1')
        df_working2=df_working[df_working['Distan1'].apply(numpy.isnan)]
        df_working2.reset_index(drop=True,inplace =True)
        #Running loop with KEY2
        df_working2=KEY_loop(df_working2, df_SQL, 'Key2', 'Distan2', 'ESPAR_ID2')
        #duplicates2=KEY_loop_replacing_uniq(df_working2, df_SQL, 'Key2', 'Distan2', 'ESPAR_ID2')
        df_working3=df_working2[df_working2['Distan2'].apply(numpy.isnan)]
        df_working3.reset_index(drop=True,inplace =True)
        #Running loop with KEY3
        df_working3=KEY_loop(df_working3, df_SQL, 'Key3', 'Distan3', 'ESPAR_ID3')
        #duplicates3=KEY_loop_replacing_uniq(df_working3, df_SQL, 'Key3', 'Distan3', 'ESPAR_ID3')


        df_working2=df_working2[['Adress1','Distan2','ESPAR_ID2']]
        #duplicates2=duplicates2[['Adress1','Distan2','ESPAR_ID2']]
        #duplicates2.columns = ['Adress1','Distan2_Dupl','ESPAR_ID2_Dupl']


        df_working3=df_working3[['Adress1','Distan3','ESPAR_ID3']]
        #duplicates3=duplicates3[['Adress1','Distan3','ESPAR_ID3']]
        #duplicates3.columns = ['Adress1','Distan3_Dupl','ESPAR_ID3_Dupl']
        #print('1')


        New = pd.merge(left=df_working,right=df_working2, how='left', left_on='Adress1', right_on='Adress1')
        #merged_left = pd.merge(left=merged_left,right=duplicates2, how='left', left_on='Adress1', right_on='Adress1')
        print('2')
        New = pd.merge(left=New,right=df_working3, how='left', left_on='Adress1', right_on='Adress1')
        #merged_left = pd.merge(left=merged_left,right=duplicates3, how='left', left_on='Adress1', right_on='Adress1')
        #print('3')


        New['Gatherer']=""
        New['Gatherer']=New['ESPAR_ID3']
        New['Gatherer']=New['Gatherer'].fillna(New['ESPAR_ID2'])


        #Distance = 0.00622166666
        Distance = 0.02
        duplicates=KEY_loop_replacing_uniq(New, df_SQL, 'Key3', 'Distan3', 'Gatherer',Distance)

        self.data_new=duplicates
    
def run():
    app = QtGui.QApplication(sys.argv)
    GUI = Window()
    sys.exit(app.exec_())


run()     
